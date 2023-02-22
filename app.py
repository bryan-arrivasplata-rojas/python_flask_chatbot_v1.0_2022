from re import A
from keras.backend import *
import nltk
nltk.download('popular')
from nltk.stem import WordNetLemmatizer
lemmatizer = WordNetLemmatizer()
import pickle
import numpy as np

from keras.models import load_model
model = load_model('modelo/model.h5')
import json
import random

import pandas as pd
import openpyxl

intents = json.loads(open('data.json').read())
words = pickle.load(open('diccionario/texts.pkl','rb'))
classes = pickle.load(open('diccionario/labels.pkl','rb'))

path_file = "resource/alumnos.xlsx"
valor = 0
contador = 0

#EXTRACCION INFORMACION NOTAS
Notas = pd.read_excel(path_file,sheet_name='Notas')
df_Notas = pd.DataFrame(Notas)
list_Codigo = df_Notas.Codigo.to_list()
list_PC1 = df_Notas.PC1.to_list()
list_PC2 = df_Notas.PC2.to_list()
list_PC3 = df_Notas.PC3.to_list()
list_PC4 = df_Notas.PC4.to_list()
list_PC5 = df_Notas.PC5.to_list()
list_PC6 = df_Notas.PC6.to_list()
list_PC7 = df_Notas.PC7.to_list()
list_Parcial1 = df_Notas.Parcial1.to_list()
list_Parcial2 = df_Notas.Parcial2.to_list()
list_EXFINAL = df_Notas.EXFINAL.to_list()
list_Promedio = df_Notas.Promedio.to_list()




Datos = pd.read_excel(path_file,sheet_name='Orden')
df_Datos = pd.DataFrame(Datos)
list_Datos = df_Datos.Codigo.to_list()
list_Orden = df_Datos.Merito.to_list()


def clean_up_sentence(sentence):
    # tokenize the pattern - split words into array
    sentence_words = nltk.word_tokenize(sentence)
    # stem each word - create short form for word
    sentence_words = [lemmatizer.lemmatize(word.lower()) for word in sentence_words]
    return sentence_words

# return bag of words array: 0 or 1 for each word in the bag that exists in the sentence

def bow(sentence, words, show_details=True):
    # tokenize the pattern
    sentence_words = clean_up_sentence(sentence)
    # bag of words - matrix of N words, vocabulary matrix
    bag = [0]*len(words)  
    for s in sentence_words:
        for i,w in enumerate(words):
            if w == s: 
                # assign 1 if current word is in the vocabulary position
                bag[i] = 1
                if show_details:
                    print ("found in bag: %s" % w)
    return(np.array(bag))

def predict_class(sentence, model):
    # filter out predictions below a threshold
    p = bow(sentence, words,show_details=False)
    res = model.predict(np.array([p]))[0]
    ERROR_THRESHOLD = 0.25
    results = [[i,r] for i,r in enumerate(res) if r>ERROR_THRESHOLD]
    # sort by strength of probability
    results.sort(key=lambda x: x[1], reverse=True)
    return_list = []
    for r in results:
        return_list.append({"intent": classes[r[0]], "probability": str(r[1])})
    return return_list

def getResponse(ints, intents_json):
    tag = ints[0]['intent']
    list_of_intents = intents_json['intents']
    #
    for i in list_of_intents:
        if(i['tag']== tag):
            result = random.choice(i['responses'])
            break
    return result

#*************
def chatbot_response(msg,valor,contador):
    if valor ==0 and contador == 0:
        ints = predict_class(msg, model)
        res = getResponse(ints, intents)
        data = interativo(msg)
        return [res,data,contador]
    elif valor ==1:
        index = list_Codigo.index(msg)
        #df = df_Notas.loc[df_Notas['Codigo'] == msg]
        #lista_df = df.apply(lambda x: x.tolist(), axis=1)
        notas_concat = extraccion_concat(index)
        return [notas_concat,0,contador]
        #return ["Tus notas actuales son: "+str(join(list_Comienzo_Notas[index])),0]
    elif valor == 2:
        index = list_Codigo.index(msg)
        #print(list_Comienzo_Promedio[index])
        return ["Tu promedio actual es: "+str(list_Promedio[index]),0,contador]
    elif valor == 3:
        indexado = list_Datos.index(msg)
        print(indexado)
        return ["Tu orden de merito actual es: "+str(list_Orden[indexado]),0,contador]
    elif valor == 4:
        #indexado = list_Datos.index(msg)
        mensaje = ""
        if contador <2:
            mensaje = "Envia la siguiente informacion solicitada : "
            contador = contador + 1
            valor == 4
            modificar_valor(msg)
        else:
            modificar_valor(msg)
            mensaje = "Datos registrados, pagar al numero de cuenta BCP 191-1234566984321 antes de los siguientes 10 dias"
            contador = 0
            valor = 0
        print(contador)
        #print(indexado)
        return [mensaje,valor,contador]

def extraccion_concat(index):
    temp=""
    temp = ("PC1: "+str(list_PC1[index])+", PC2:"+str(list_PC2[index])+", PC3:"+str(list_PC3[index])+", PC4:"+str(list_PC4[index])+
                ", PC5:"+str(list_PC5[index])+", PC6:"+str(list_PC6[index])+", PC7:"+str(list_PC7[index])+", Parcial 1:"+str(list_Parcial1[index])+
                ", Parcial 2:"+str(list_Parcial2[index])+", Ex.Final:"+str(list_EXFINAL[index]))
    return temp

def agregar_fila_excel():
    Reservacion = pd.read_excel(path_file,sheet_name='PreMatricula')
    df_Reservacion = pd.DataFrame(Reservacion)
    list_Reservacion_Identificador = df_Reservacion.Identificador.to_list()
    reserva_contador_Identificador = len(list_Reservacion_Identificador)
    wbkName = 'resource/alumnos.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    wks = wbk['PreMatricula']
    wks.append([reserva_contador_Identificador+1])
    wbk.save(wbkName)
    wbk.close

def modificar_valor(informacion_recibida):
    Reservacion = pd.read_excel(path_file,sheet_name='PreMatricula')
    df_Reservacion = pd.DataFrame(Reservacion)
    list_Reservacion_Identificador = df_Reservacion.Identificador.to_list()
    list_Reservacion_Nombres = df_Reservacion.NombreCompleto.to_list()
    list_Reservacion_DNI = df_Reservacion.DNI.to_list()
    list_Reservacion_Celular = df_Reservacion.Celular.to_list()
    reserva_contador_Identificador = len(list_Reservacion_Identificador)
    wbkName = 'resource/alumnos.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    wks = wbk['PreMatricula']
    #print(list_Reservacion_Nombres[len(list_Reservacion_Nombres)-1])

    if pd.isna(list_Reservacion_Nombres[len(list_Reservacion_Nombres)-1]):
        wks.cell(row=reserva_contador_Identificador+1, column=2).value = informacion_recibida
        wbk.save(wbkName)
        wbk.close
        return True
    elif pd.isna(list_Reservacion_DNI[len(list_Reservacion_DNI)-1]):
        wks.cell(row=reserva_contador_Identificador+1, column=3).value = informacion_recibida
        wbk.save(wbkName)
        wbk.close
        return True
    elif pd.isna(list_Reservacion_Celular[len(list_Reservacion_DNI)-1]):
        wks.cell(row=reserva_contador_Identificador+1, column=4).value = informacion_recibida
        wbk.save(wbkName)
        wbk.close
        return True
    else:
        return False

def interativo(palabra):
    if palabra.__contains__(str("notas")):
        print("Vas a ver tus notas")
        return 1
    elif palabra.__contains__(str("promedio")):
        print("Vas a ver tu promedio")
        return 2
    elif palabra.__contains__(str("orden")) or palabra.__contains__(str("posicion")) and palabra.__contains__(str("puesto")):
        print("Vas a ver tu orden")
        return 3
    elif palabra.__contains__(str("matricula")):
        print("Vas a reservar tu matricula")
        agregar_fila_excel()
        return 4
    elif palabra.__contains__(str("2017")):
        print("Tus notas son")
        return 99
    else:
        return 0

from flask import Flask, render_template, request

app = Flask(__name__)
app.static_folder = 'static'

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/get")
def get_bot_response():
    global valor
    global contador

    userText = request.args.get('msg')

    
    #valor = interativo(userText)

    data = chatbot_response(userText,valor,contador)

    valor = data[1]

    contador = data[2]
    
    return data[0]


if __name__ == "__main__":
    app.run(debug=True)